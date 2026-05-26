from __future__ import annotations

import argparse
import os
import smtplib
from dataclasses import dataclass
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REQUIRED_COLUMNS = {
    "order_id",
    "order_date",
    "customer",
    "region",
    "category",
    "product",
    "quantity",
    "unit_price",
    "discount",
}


@dataclass(frozen=True)
class EmailConfig:
    sender: str
    password: str
    recipients: list[str]
    smtp_host: str = "smtp.gmail.com"
    smtp_port: int = 587


def load_and_clean_sales(csv_path: Path) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    missing = REQUIRED_COLUMNS.difference(df.columns)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    df = df.drop_duplicates(subset=["order_id"]).copy()
    df["order_date"] = parse_order_dates(df["order_date"])
    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce").fillna(0).astype(int)
    df["unit_price"] = pd.to_numeric(df["unit_price"], errors="coerce").fillna(0.0)
    df["discount"] = pd.to_numeric(df["discount"], errors="coerce").fillna(0.0).clip(0, 1)

    text_columns = ["customer", "region", "category", "product"]
    for column in text_columns:
        df[column] = df[column].astype(str).str.strip().str.title()

    df = df[df["order_date"].notna()]
    df = df[df["quantity"] > 0]
    df = df[df["unit_price"] >= 0]
    df["gross_revenue"] = df["quantity"] * df["unit_price"]
    df["net_revenue"] = df["gross_revenue"] * (1 - df["discount"])
    df["month"] = df["order_date"].dt.to_period("M").astype(str)
    df["quarter"] = df["order_date"].dt.to_period("Q").astype(str)
    return df.sort_values("order_date").reset_index(drop=True)


def parse_order_dates(values: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(values, errors="coerce", format="%Y-%m-%d")
    missing = parsed.isna()
    if missing.any():
        parsed.loc[missing] = pd.to_datetime(
            values.loc[missing],
            errors="coerce",
            dayfirst=True,
        )
    return parsed


def build_analysis(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    by_region = summarize_dimension(df, "region")
    by_category = summarize_dimension(df, "category")
    by_product = summarize_dimension(df, "product").head(10)
    monthly = (
        df.groupby("month", as_index=False)
        .agg(
            orders=("order_id", "nunique"),
            units_sold=("quantity", "sum"),
            net_revenue=("net_revenue", "sum"),
            avg_order_value=("net_revenue", "mean"),
        )
        .sort_values("month")
    )

    return {
        "Cleaned Data": df,
        "Revenue By Region": by_region,
        "Revenue By Category": by_category,
        "Top Products": by_product,
        "Monthly Trend": monthly,
    }


def summarize_dimension(df: pd.DataFrame, dimension: str) -> pd.DataFrame:
    return (
        df.groupby(dimension, as_index=False)
        .agg(
            orders=("order_id", "nunique"),
            units_sold=("quantity", "sum"),
            gross_revenue=("gross_revenue", "sum"),
            net_revenue=("net_revenue", "sum"),
            avg_discount=("discount", "mean"),
        )
        .sort_values("net_revenue", ascending=False)
    )


def write_excel_report(analysis: dict[str, pd.DataFrame], output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, data in analysis.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)

        kpis = calculate_kpis(analysis["Cleaned Data"])
        pd.DataFrame(kpis.items(), columns=["Metric", "Value"]).to_excel(
            writer, sheet_name="KPI Summary", index=False
        )

    workbook = load_workbook(output_path)
    for worksheet in workbook.worksheets:
        format_sheet(worksheet)

    add_charts(workbook)
    workbook.save(output_path)
    return output_path


def calculate_kpis(df: pd.DataFrame) -> dict[str, object]:
    top_region = df.groupby("region")["net_revenue"].sum().idxmax()
    top_category = df.groupby("category")["net_revenue"].sum().idxmax()
    return {
        "Report Generated At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Total Orders": int(df["order_id"].nunique()),
        "Units Sold": int(df["quantity"].sum()),
        "Gross Revenue": round(float(df["gross_revenue"].sum()), 2),
        "Net Revenue": round(float(df["net_revenue"].sum()), 2),
        "Average Order Value": round(float(df["net_revenue"].mean()), 2),
        "Top Region": top_region,
        "Top Category": top_category,
    }


def format_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_length + 2, 12), 34)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00'


def add_charts(workbook) -> None:
    add_bar_chart(
        workbook["Revenue By Region"],
        title="Net Revenue by Region",
        anchor="H2",
        category_col=1,
        value_col=4,
    )
    add_bar_chart(
        workbook["Revenue By Category"],
        title="Net Revenue by Category",
        anchor="H2",
        category_col=1,
        value_col=4,
    )
    add_line_chart(workbook["Monthly Trend"], title="Monthly Revenue Trend", anchor="G2")


def add_bar_chart(worksheet, title: str, anchor: str, category_col: int, value_col: int) -> None:
    max_row = worksheet.max_row
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "Net Revenue"
    chart.x_axis.title = worksheet.cell(row=1, column=category_col).value
    data = Reference(worksheet, min_col=value_col, min_row=1, max_row=max_row)
    categories = Reference(worksheet, min_col=category_col, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 12
    worksheet.add_chart(chart, anchor)


def add_line_chart(worksheet, title: str, anchor: str) -> None:
    max_row = worksheet.max_row
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "Net Revenue"
    chart.x_axis.title = "Month"
    data = Reference(worksheet, min_col=4, min_row=1, max_row=max_row)
    categories = Reference(worksheet, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 7
    chart.width = 12
    worksheet.add_chart(chart, anchor)


def load_email_config() -> EmailConfig | None:
    sender = os.getenv("SMTP_SENDER")
    password = os.getenv("SMTP_PASSWORD")
    recipients = os.getenv("SMTP_RECIPIENTS", "")
    if not sender or not password or not recipients:
        return None

    return EmailConfig(
        sender=sender,
        password=password,
        recipients=[email.strip() for email in recipients.split(",") if email.strip()],
        smtp_host=os.getenv("SMTP_HOST", "smtp.gmail.com"),
        smtp_port=int(os.getenv("SMTP_PORT", "587")),
    )


def email_report(report_path: Path, config: EmailConfig) -> None:
    message = EmailMessage()
    message["Subject"] = f"Automated Sales Report - {datetime.now():%Y-%m-%d}"
    message["From"] = config.sender
    message["To"] = ", ".join(config.recipients)
    message.set_content(
        "Hello,\n\nThe latest automated sales report is attached.\n\nRegards,\nSales Automation Bot"
    )

    message.add_attachment(
        report_path.read_bytes(),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=report_path.name,
    )

    with smtplib.SMTP(config.smtp_host, config.smtp_port) as smtp:
        smtp.starttls()
        smtp.login(config.sender, config.password)
        smtp.send_message(message)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate an automated sales Excel report.")
    parser.add_argument("--input", type=Path, default=Path("sample_sales.csv"), help="Raw sales CSV path.")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("output") / "sales_report.xlsx",
        help="Generated Excel report path.",
    )
    parser.add_argument(
        "--email",
        action="store_true",
        help="Email the report using SMTP_* environment variables.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    sales = load_and_clean_sales(args.input)
    analysis = build_analysis(sales)
    report_path = write_excel_report(analysis, args.output)
    print(f"Report generated: {report_path.resolve()}")

    if args.email:
        config = load_email_config()
        if config is None:
            raise RuntimeError(
                "Email requested, but SMTP_SENDER, SMTP_PASSWORD, and SMTP_RECIPIENTS are not set."
            )
        email_report(report_path, config)
        print(f"Report emailed to: {', '.join(config.recipients)}")


if __name__ == "__main__":
    main()
